import io

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def build_pdf(summary: dict, timeline: dict, meta: dict) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Scoutia — Reporte de Leads", styles["Heading1"]))
    elements.append(Paragraph(f"Workspace: {meta.get('workspace_name', '')} | Generado por: {meta.get('full_name', '')}", styles["Normal"]))
    elements.append(Paragraph(f"Fecha: {meta.get('generated_date', '')} | Período: {meta.get('period_label', '')}", styles["Normal"]))
    elements.append(Spacer(1, 0.5*cm))

    total = summary.get("total_leads", 0)
    contacted = summary.get("contacted", 0)
    qualified = summary.get("by_status", {}).get("calificado", 0)
    conversion = round(qualified / total * 100) if total else 0
    kpi_data = [
        ["Total Leads", "Contactados", "Conversión", "Alta oportunidad"],
        [str(total), str(contacted), f"{conversion}%", str(meta.get("high_opportunity", 0))],
    ]
    kpi_table = Table(kpi_data, colWidths=[4*cm, 4*cm, 4*cm, 4*cm])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17110D")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("BOX", (0, 0), (-1, -1), 1, colors.black),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 0.4*cm))

    elements.append(Paragraph("Pipeline por etapa", styles["Heading3"]))
    status_labels = {"nuevo": "Nuevos", "contactado": "Contactados", "calificado": "Calificados", "perdido": "Perdidos", "desvinculado": "Desvinculados"}
    by_status = summary.get("by_status", {})
    status_data = [["Etapa", "Cantidad", "%"]] + [
        [status_labels.get(s, s), str(by_status.get(s, 0)), f"{round(by_status.get(s, 0) / total * 100) if total else 0}%"]
        for s in ["nuevo", "contactado", "calificado", "perdido", "desvinculado"]
    ]
    status_table = Table(status_data, colWidths=[6*cm, 4*cm, 3*cm])
    status_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#333333")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOX", (0, 0), (-1, -1), 1, colors.black),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
    ]))
    elements.append(status_table)
    elements.append(Spacer(1, 0.4*cm))

    elements.append(Paragraph("Por prioridad", styles["Heading3"]))
    by_priority = summary.get("by_priority", {})
    priority_labels = {"alta": "Alta", "media": "Media", "baja": "Baja"}
    priority_data = [["Prioridad", "Cantidad"]] + [
        [priority_labels.get(p, p), str(by_priority.get(p, 0))]
        for p in ["alta", "media", "baja"]
    ]
    priority_table = Table(priority_data, colWidths=[6*cm, 4*cm])
    priority_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#333333")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOX", (0, 0), (-1, -1), 1, colors.black),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
    ]))
    elements.append(priority_table)
    elements.append(Spacer(1, 0.4*cm))

    elements.append(Paragraph("Categorías principales", styles["Heading3"]))
    by_category = summary.get("by_category", {})
    sorted_cats = sorted(by_category.items(), key=lambda x: x[1], reverse=True)
    top_cats = sorted_cats[:8]
    other = sum(v for _, v in sorted_cats[8:])
    cat_rows = [[cat, str(cnt)] for cat, cnt in top_cats]
    if other:
        cat_rows.append(["Otras", str(other)])
    cat_data = [["Categoría", "Cantidad"]] + cat_rows
    cat_table = Table(cat_data, colWidths=[9*cm, 4*cm])
    cat_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#333333")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOX", (0, 0), (-1, -1), 1, colors.black),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
    ]))
    elements.append(cat_table)
    elements.append(Spacer(1, 0.4*cm))

    points = timeline.get("points", [])
    period_total = sum(p["leads"] for p in points)
    best_day = max(points, key=lambda p: p["leads"], default=None)
    best_str = f"{best_day['date']}: {best_day['leads']} leads" if best_day and best_day["leads"] > 0 else "Sin actividad"
    elements.append(Paragraph(f"Actividad en el período: {period_total} leads nuevos. Mejor día: {best_str}", styles["Normal"]))
    elements.append(Spacer(1, 0.5*cm))

    elements.append(Paragraph("scoutia.dev | Reporte generado automáticamente", styles["Normal"]))

    doc.build(elements)
    return buf.getvalue()


def build_xlsx(summary: dict, timeline: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Resumen"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="17110D")

    def write_section_header(ws, row, text):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(bold=True)
        return row + 1

    def write_table(ws, row, headers, rows_data):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")
        row += 1
        for r in rows_data:
            for col, val in enumerate(r, 1):
                ws.cell(row=row, column=col, value=val)
            row += 1
        return row + 1

    total = summary.get("total_leads", 0)
    contacted = summary.get("contacted", 0)
    avg_score = summary.get("avg_score", 0)
    qualified = summary.get("by_status", {}).get("calificado", 0)
    conversion = round(qualified / total * 100) if total else 0

    row = 1
    row = write_section_header(ws1, row, "KPIs")
    row = write_table(ws1, row, ["Métrica", "Valor"], [
        ["Total Leads", total],
        ["Contactados", contacted],
        ["Conversión %", f"{conversion}%"],
        ["Score Promedio", avg_score],
    ])

    by_status = summary.get("by_status", {})
    status_labels = {"nuevo": "Nuevos", "contactado": "Contactados", "calificado": "Calificados", "perdido": "Perdidos", "desvinculado": "Desvinculados"}
    row = write_section_header(ws1, row, "Por Estado")
    row = write_table(ws1, row, ["Estado", "Cantidad", "%"], [
        [status_labels.get(s, s), by_status.get(s, 0), f"{round(by_status.get(s, 0) / total * 100) if total else 0}%"]
        for s in ["nuevo", "contactado", "calificado", "perdido", "desvinculado"]
    ])

    by_priority = summary.get("by_priority", {})
    priority_labels = {"alta": "Alta", "media": "Media", "baja": "Baja"}
    row = write_section_header(ws1, row, "Por Prioridad")
    row = write_table(ws1, row, ["Prioridad", "Cantidad"], [
        [priority_labels.get(p, p), by_priority.get(p, 0)]
        for p in ["alta", "media", "baja"]
    ])

    by_category = summary.get("by_category", {})
    sorted_cats = sorted(by_category.items(), key=lambda x: x[1], reverse=True)
    row = write_section_header(ws1, row, "Por Categoría")
    row = write_table(ws1, row, ["Categoría", "Cantidad"], [[cat, cnt] for cat, cnt in sorted_cats])

    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 15
    ws1.column_dimensions["C"].width = 10

    ws2 = wb.create_sheet("Actividad")
    ws2.cell(row=1, column=1, value="Fecha").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2, value="Leads").font = header_font
    ws2.cell(row=1, column=2).fill = header_fill
    ws2.freeze_panes = "A2"
    for i, point in enumerate(timeline.get("points", []), start=2):
        ws2.cell(row=i, column=1, value=point["date"])
        ws2.cell(row=i, column=2, value=point["leads"])
    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 10

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
